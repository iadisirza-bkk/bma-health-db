"""Excel parser for district-level health screening data.

Parses .xlsx files from the Medical Service Department into the
district_health_data.json structure used by the data adapter layer.

Sync port from the source project.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column mapping: Thai header -> (disease_key, indicator_key, field)
# ---------------------------------------------------------------------------

# Core columns (always required)
CORE_COLUMNS: dict[str, str] = {
    "รหัสเขต": "dcode",
    "ชื่อเขต": "name_th",
    "ชื่อเขต_en": "name_en",
    "จำนวนคัดกรอง": "total_screened",
}

# Disease-level percentage columns
DISEASE_PCT_COLUMNS: dict[str, tuple[str, str, str]] = {
    # header -> (disease_key, thai_name, english_name)
    "เบาหวาน_pct": ("diabetes", "เบาหวาน", "Diabetes"),
    "ความดัน_pct": ("hypertension", "ความดันโลหิตสูง", "Hypertension"),
    "หลอดเลือดหัวใจ_pct": ("cardiovascular", "โรคหลอดเลือดหัวใจ", "Cardiovascular"),
    "โรคอ้วน_pct": ("obesity", "โรคอ้วน", "Obesity"),
    "ไขมัน_pct": ("dyslipidemia", "ไขมันในเลือดผิดปกติ", "Dyslipidemia"),
}

# Detailed indicator columns (optional)
# Format: header -> (disease_key, indicator_key, field_name)
INDICATOR_COLUMNS: dict[str, tuple[str, str, str]] = {
    # Diabetes indicators
    "FBS_mean": ("diabetes", "fbs", "mean"),
    "FBS_cutoff": ("diabetes", "fbs", "pct_above_cutoff"),
    "FBS_count_above": ("diabetes", "fbs", "count_above"),
    "FBS_pre_pct": ("diabetes", "fbs", "pct_pre_cutoff"),
    "FBS_pre_count": ("diabetes", "fbs", "count_pre"),
    "BMI_diabetes_mean": ("diabetes", "bmi", "mean"),
    "BMI_diabetes_cutoff": ("diabetes", "bmi", "pct_above_cutoff"),
    "BMI_diabetes_count_above": ("diabetes", "bmi", "count_above"),
    "BMI_diabetes_pre_pct": ("diabetes", "bmi", "pct_pre_cutoff"),
    "BMI_diabetes_pre_count": ("diabetes", "bmi", "count_pre"),
    "WAIST_diabetes_mean": ("diabetes", "waist", "mean"),
    "WAIST_diabetes_cutoff": ("diabetes", "waist", "pct_above_cutoff"),
    "WAIST_diabetes_count_above": ("diabetes", "waist", "count_above"),
    # Hypertension indicators
    "SBP_mean": ("hypertension", "sbp", "mean"),
    "SBP_cutoff": ("hypertension", "sbp", "pct_above_cutoff"),
    "SBP_count_above": ("hypertension", "sbp", "count_above"),
    "SBP_pre_pct": ("hypertension", "sbp", "pct_pre_cutoff"),
    "SBP_pre_count": ("hypertension", "sbp", "count_pre"),
    "DBP_mean": ("hypertension", "dbp", "mean"),
    "DBP_cutoff": ("hypertension", "dbp", "pct_above_cutoff"),
    "DBP_count_above": ("hypertension", "dbp", "count_above"),
    "DBP_pre_pct": ("hypertension", "dbp", "pct_pre_cutoff"),
    "DBP_pre_count": ("hypertension", "dbp", "count_pre"),
    # Cardiovascular indicators
    "CHOL_mean": ("cardiovascular", "cholesterol", "mean"),
    "CHOL_cutoff": ("cardiovascular", "cholesterol", "pct_above_cutoff"),
    "CHOL_count_above": ("cardiovascular", "cholesterol", "count_above"),
    "CHOL_pre_pct": ("cardiovascular", "cholesterol", "pct_pre_cutoff"),
    "CHOL_pre_count": ("cardiovascular", "cholesterol", "count_pre"),
    "LDL_cardio_mean": ("cardiovascular", "ldl", "mean"),
    "LDL_cardio_cutoff": ("cardiovascular", "ldl", "pct_above_cutoff"),
    "LDL_cardio_count_above": ("cardiovascular", "ldl", "count_above"),
    "LDL_cardio_pre_pct": ("cardiovascular", "ldl", "pct_pre_cutoff"),
    "LDL_cardio_pre_count": ("cardiovascular", "ldl", "count_pre"),
    "HDL_mean": ("cardiovascular", "hdl", "mean"),
    "HDL_cutoff": ("cardiovascular", "hdl", "pct_below_cutoff"),
    "HDL_count_below": ("cardiovascular", "hdl", "count_below"),
    "TG_cardio_mean": ("cardiovascular", "triglyceride", "mean"),
    "TG_cardio_cutoff": ("cardiovascular", "triglyceride", "pct_above_cutoff"),
    "TG_cardio_count_above": ("cardiovascular", "triglyceride", "count_above"),
    "TG_cardio_pre_pct": ("cardiovascular", "triglyceride", "pct_pre_cutoff"),
    "TG_cardio_pre_count": ("cardiovascular", "triglyceride", "count_pre"),
    # Obesity indicators
    "BMI_obesity_mean": ("obesity", "bmi", "mean"),
    "BMI_obesity_cutoff": ("obesity", "bmi", "pct_above_cutoff"),
    "BMI_obesity_count_above": ("obesity", "bmi", "count_above"),
    "BMI_obesity_pre_pct": ("obesity", "bmi", "pct_pre_cutoff"),
    "BMI_obesity_pre_count": ("obesity", "bmi", "count_pre"),
    "WAIST_obesity_mean": ("obesity", "waist", "mean"),
    "WAIST_obesity_cutoff": ("obesity", "waist", "pct_above_cutoff"),
    "WAIST_obesity_count_above": ("obesity", "waist", "count_above"),
    # Dyslipidemia indicators
    "CHOL_dyslip_mean": ("dyslipidemia", "cholesterol", "mean"),
    "CHOL_dyslip_cutoff": ("dyslipidemia", "cholesterol", "pct_above_cutoff"),
    "CHOL_dyslip_count_above": ("dyslipidemia", "cholesterol", "count_above"),
    "TG_dyslip_mean": ("dyslipidemia", "triglyceride", "mean"),
    "TG_dyslip_cutoff": ("dyslipidemia", "triglyceride", "pct_above_cutoff"),
    "TG_dyslip_count_above": ("dyslipidemia", "triglyceride", "count_above"),
    "LDL_dyslip_mean": ("dyslipidemia", "ldl", "mean"),
    "LDL_dyslip_cutoff": ("dyslipidemia", "ldl", "pct_above_cutoff"),
    "LDL_dyslip_count_above": ("dyslipidemia", "ldl", "count_above"),
    "HDL_dyslip_mean": ("dyslipidemia", "hdl", "mean"),
    "HDL_dyslip_cutoff": ("dyslipidemia", "hdl", "pct_below_cutoff"),
    "HDL_dyslip_count_below": ("dyslipidemia", "hdl", "count_below"),
}

# ---------------------------------------------------------------------------
# Static indicator metadata (zones, labels, units, cutoffs, bar_max)
# ---------------------------------------------------------------------------

INDICATOR_META: dict[str, dict[str, Any]] = {
    "fbs": {
        "label": "น้ำตาลในเลือด (FBS)", "unit": "mg/dL", "bar_max": 200,
        "cutoff": 126, "cutoff_pre": 100,
        "zones": [{"label": "ปกติ", "max": 100, "color": "#22c55e"}, {"label": "เสี่ยง", "max": 126, "color": "#eab308"}, {"label": "เบาหวาน", "max": 200, "color": "#ef4444"}],
    },
    "bmi": {
        "label": "ดัชนีมวลกาย (BMI)", "unit": "kg/m2", "bar_max": 40,
        "cutoff": 25, "cutoff_pre": 23,
        "zones": [{"label": "ปกติ", "max": 23, "color": "#22c55e"}, {"label": "น้ำหนักเกิน", "max": 25, "color": "#eab308"}, {"label": "อ้วน", "max": 40, "color": "#ef4444"}],
    },
    "bmi_obesity": {
        "label": "ดัชนีมวลกาย (BMI)", "unit": "kg/m2", "bar_max": 40,
        "cutoff": 25, "cutoff_pre": 23,
        "zones": [{"label": "ผอม", "max": 18.5, "color": "#3b82f6"}, {"label": "ปกติ", "max": 23, "color": "#22c55e"}, {"label": "น้ำหนักเกิน", "max": 25, "color": "#eab308"}, {"label": "อ้วน", "max": 40, "color": "#ef4444"}],
    },
    "waist": {
        "label": "รอบเอว", "unit": "cm", "bar_max": 130, "cutoff": 85,
        "zones": [{"label": "ปกติ", "max": 85, "color": "#22c55e"}, {"label": "เสี่ยง", "max": 130, "color": "#ef4444"}],
    },
    "waist_obesity": {
        "label": "รอบเอว", "unit": "cm", "bar_max": 130, "cutoff": 85,
        "zones": [{"label": "ปกติ", "max": 85, "color": "#22c55e"}, {"label": "อ้วนลงพุง", "max": 130, "color": "#ef4444"}],
    },
    "sbp": {
        "label": "ความดันตัวบน (SBP)", "unit": "mmHg", "bar_max": 200,
        "cutoff": 140, "cutoff_pre": 130,
        "zones": [{"label": "ปกติ", "max": 130, "color": "#22c55e"}, {"label": "เริ่มสูง", "max": 140, "color": "#eab308"}, {"label": "ความดันสูง", "max": 200, "color": "#ef4444"}],
    },
    "dbp": {
        "label": "ความดันตัวล่าง (DBP)", "unit": "mmHg", "bar_max": 130,
        "cutoff": 90, "cutoff_pre": 85,
        "zones": [{"label": "ปกติ", "max": 85, "color": "#22c55e"}, {"label": "เริ่มสูง", "max": 90, "color": "#eab308"}, {"label": "ความดันสูง", "max": 130, "color": "#ef4444"}],
    },
    "cholesterol": {
        "label": "คอเลสเตอรอลรวม", "unit": "mg/dL", "bar_max": 350,
        "cutoff": 240, "cutoff_pre": 200,
        "zones": [{"label": "ปกติ", "max": 200, "color": "#22c55e"}, {"label": "สูงปานกลาง", "max": 240, "color": "#eab308"}, {"label": "สูง", "max": 350, "color": "#ef4444"}],
    },
    "cholesterol_dyslip": {
        "label": "คอเลสเตอรอลรวม", "unit": "mg/dL", "bar_max": 350, "cutoff": 200,
        "zones": [{"label": "ปกติ", "max": 200, "color": "#22c55e"}, {"label": "สูง", "max": 350, "color": "#ef4444"}],
    },
    "ldl": {
        "label": "LDL (ไขมันไม่ดี)", "unit": "mg/dL", "bar_max": 250,
        "cutoff": 160, "cutoff_pre": 130,
        "zones": [{"label": "ปกติ", "max": 130, "color": "#22c55e"}, {"label": "สูงปานกลาง", "max": 160, "color": "#eab308"}, {"label": "สูง", "max": 250, "color": "#ef4444"}],
    },
    "ldl_dyslip": {
        "label": "LDL (ไขมันไม่ดี)", "unit": "mg/dL", "bar_max": 250,
        "cutoff": 160, "cutoff_pre": 130,
        "zones": [{"label": "ปกติ", "max": 130, "color": "#22c55e"}, {"label": "สูงปานกลาง", "max": 160, "color": "#eab308"}, {"label": "สูง", "max": 250, "color": "#ef4444"}],
    },
    "hdl": {
        "label": "HDL (ไขมันดี)", "unit": "mg/dL", "bar_max": 100,
        "cutoff": 40, "direction": "below",
        "zones": [{"label": "ต่ำ (เสี่ยง)", "max": 40, "color": "#ef4444"}, {"label": "ปกติ", "max": 100, "color": "#22c55e"}],
    },
    "triglyceride": {
        "label": "ไตรกลีเซอไรด์", "unit": "mg/dL", "bar_max": 400,
        "cutoff": 200, "cutoff_pre": 150,
        "zones": [{"label": "ปกติ", "max": 150, "color": "#22c55e"}, {"label": "สูงปานกลาง", "max": 200, "color": "#eab308"}, {"label": "สูง", "max": 400, "color": "#ef4444"}],
    },
    "triglyceride_dyslip": {
        "label": "ไตรกลีเซอไรด์", "unit": "mg/dL", "bar_max": 400, "cutoff": 150,
        "zones": [{"label": "ปกติ", "max": 150, "color": "#22c55e"}, {"label": "สูง", "max": 400, "color": "#ef4444"}],
    },
}

_INDICATOR_META_KEY: dict[tuple[str, str], str] = {
    ("diabetes", "fbs"): "fbs",
    ("diabetes", "bmi"): "bmi",
    ("diabetes", "waist"): "waist",
    ("hypertension", "sbp"): "sbp",
    ("hypertension", "dbp"): "dbp",
    ("cardiovascular", "cholesterol"): "cholesterol",
    ("cardiovascular", "ldl"): "ldl",
    ("cardiovascular", "hdl"): "hdl",
    ("cardiovascular", "triglyceride"): "triglyceride",
    ("obesity", "bmi"): "bmi_obesity",
    ("obesity", "waist"): "waist_obesity",
    ("dyslipidemia", "cholesterol"): "cholesterol_dyslip",
    ("dyslipidemia", "triglyceride"): "triglyceride_dyslip",
    ("dyslipidemia", "ldl"): "ldl_dyslip",
    ("dyslipidemia", "hdl"): "hdl",
}


# ---------------------------------------------------------------------------
# Parse result
# ---------------------------------------------------------------------------


@dataclass
class ParseResult:
    """Result of parsing an Excel file."""

    data: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.data) > 0 and len(self.errors) == 0


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def _safe_float(value: Any, row: int, col_name: str, errors: list[str]) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        errors.append(f"Row {row}: column '{col_name}' has non-numeric value '{value}'")
        return None


def _safe_int(value: Any, row: int, col_name: str, errors: list[str]) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        errors.append(f"Row {row}: column '{col_name}' has non-numeric value '{value}'")
        return None


def _build_indicator(
    disease_key: str,
    indicator_key: str,
    values: dict[str, Any],
    total_screened: int,
) -> dict[str, Any]:
    """Build a full indicator dict from parsed values and static metadata."""
    meta_key = _INDICATOR_META_KEY.get((disease_key, indicator_key), indicator_key)
    meta = INDICATOR_META.get(meta_key, {})

    indicator: dict[str, Any] = {
        "label": meta.get("label", indicator_key),
        "unit": meta.get("unit", ""),
        "bar_max": meta.get("bar_max", 100),
        "mean": values.get("mean", 0.0),
        "cutoff": meta.get("cutoff", 0),
        "zones": meta.get("zones", []),
    }

    if meta.get("direction") == "below":
        indicator["direction"] = "below"
        if "pct_below_cutoff" in values:
            indicator["pct_below_cutoff"] = values["pct_below_cutoff"]
        if "count_below" in values:
            indicator["count_below"] = values["count_below"]
    else:
        if "pct_above_cutoff" in values:
            indicator["pct_above_cutoff"] = values["pct_above_cutoff"]
        if "count_above" in values:
            indicator["count_above"] = values["count_above"]

    if meta.get("cutoff_pre") is not None:
        indicator["cutoff_pre"] = meta["cutoff_pre"]
    if "pct_pre_cutoff" in values:
        indicator["pct_pre_cutoff"] = values["pct_pre_cutoff"]
    if "count_pre" in values:
        indicator["count_pre"] = values["count_pre"]

    return indicator


def parse_health_excel(file_bytes: bytes) -> ParseResult:
    """Parse an .xlsx file containing district health screening data.

    Parameters
    ----------
    file_bytes:
        Raw bytes of the uploaded .xlsx file.

    Returns
    -------
    ParseResult with ``data`` matching the district_health_data.json structure
    and any row-level ``errors``.
    """
    result = ParseResult()

    try:
        wb: Workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        result.errors.append(f"Cannot open Excel file: {exc}")
        return result

    ws: Worksheet = wb.active  # type: ignore[assignment]
    if ws is None:
        result.errors.append("Workbook has no active sheet")
        return result

    # --- Read header row ---
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        result.errors.append("Sheet has no header row")
        return result

    headers: list[str] = [str(h).strip() if h is not None else "" for h in header_row]

    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            col_idx[h] = i

    missing_core = [h for h in CORE_COLUMNS if h not in col_idx]
    if missing_core:
        result.errors.append(f"Missing required columns: {', '.join(missing_core)}")
        return result

    present_diseases = {h: v for h, v in DISEASE_PCT_COLUMNS.items() if h in col_idx}
    if not present_diseases:
        result.errors.append(
            "No disease percentage columns found. "
            f"Expected at least one of: {', '.join(DISEASE_PCT_COLUMNS.keys())}"
        )
        return result

    # --- Parse data rows ---
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue

        cells = list(row)

        def _cell(col_name: str) -> Any:
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(cells):
                return None
            return cells[idx]

        dcode_raw = _cell("รหัสเขต")
        if dcode_raw is None or str(dcode_raw).strip() == "":
            continue

        dcode = str(int(float(dcode_raw))) if isinstance(dcode_raw, (int, float)) else str(dcode_raw).strip()
        name_th = str(_cell("ชื่อเขต") or "").strip()
        name_en = str(_cell("ชื่อเขต_en") or "").strip()
        total_screened = _safe_int(_cell("จำนวนคัดกรอง"), row_num, "จำนวนคัดกรอง", result.errors)

        if not name_th:
            result.errors.append(f"Row {row_num} (dcode={dcode}): missing ชื่อเขต")
            continue
        if total_screened is None or total_screened <= 0:
            result.errors.append(f"Row {row_num} (dcode={dcode}): invalid จำนวนคัดกรอง")
            continue

        diseases: dict[str, Any] = {}

        for header, (disease_key, thai_name, eng_name) in present_diseases.items():
            pct = _safe_float(_cell(header), row_num, header, result.errors)
            if pct is None:
                continue

            disease_entry: dict[str, Any] = {
                "name": thai_name,
                "name_en": eng_name,
                "pct_at_risk": round(pct, 1),
                "total_screened": total_screened,
                "indicators": {},
            }

            indicator_values: dict[str, dict[str, Any]] = {}
            for ind_header, (d_key, ind_key, field_name) in INDICATOR_COLUMNS.items():
                if d_key != disease_key:
                    continue
                if ind_header not in col_idx:
                    continue
                val = _cell(ind_header)
                if val is None:
                    continue

                if ind_key not in indicator_values:
                    indicator_values[ind_key] = {}

                parsed = _safe_float(val, row_num, ind_header, result.errors)
                if parsed is not None:
                    if "count" in field_name:
                        indicator_values[ind_key][field_name] = int(parsed)
                    else:
                        indicator_values[ind_key][field_name] = round(parsed, 1)

            for ind_key, values in indicator_values.items():
                disease_entry["indicators"][ind_key] = _build_indicator(
                    disease_key, ind_key, values, total_screened
                )

            diseases[disease_key] = disease_entry

        if not diseases:
            result.errors.append(f"Row {row_num} (dcode={dcode}): no valid disease data")
            continue

        result.data[dcode] = {
            "dcode": dcode,
            "name_th": name_th,
            "name_en": name_en,
            "total_screened": total_screened,
            "diseases": diseases,
        }

    if not result.data and not result.errors:
        result.errors.append("No data rows found in the sheet")

    try:
        wb.close()
    except Exception:
        pass

    logger.info("Parsed Excel: %d districts, %d errors", len(result.data), len(result.errors))
    return result


# ---------------------------------------------------------------------------
# Template generator
# ---------------------------------------------------------------------------


def generate_template_workbook() -> bytes:
    """Generate a template .xlsx with correct headers and one example row.

    Returns raw bytes of the workbook.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "ข้อมูลคัดกรองสุขภาพ"

    headers: list[str] = list(CORE_COLUMNS.keys())
    headers.extend(DISEASE_PCT_COLUMNS.keys())
    headers.extend(INDICATOR_COLUMNS.keys())

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)

    example: dict[str, Any] = {
        "รหัสเขต": 1001,
        "ชื่อเขต": "พระนคร",
        "ชื่อเขต_en": "Phra Nakhon",
        "จำนวนคัดกรอง": 25952,
        "เบาหวาน_pct": 21.8,
        "ความดัน_pct": 28.9,
        "หลอดเลือดหัวใจ_pct": 17.5,
        "โรคอ้วน_pct": 38.1,
        "ไขมัน_pct": 36.9,
        "FBS_mean": 92.2,
        "FBS_cutoff": 6.1,
        "FBS_count_above": 1586,
        "FBS_pre_pct": 29.7,
        "FBS_pre_count": 7709,
        "SBP_mean": 131.8,
        "SBP_cutoff": 32.4,
        "SBP_count_above": 8410,
        "SBP_pre_pct": 21.7,
        "SBP_pre_count": 5625,
        "DBP_mean": 82.1,
        "DBP_cutoff": 23.7,
        "DBP_count_above": 6150,
        "DBP_pre_pct": 15.7,
        "DBP_pre_count": 4063,
    }

    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=example.get(header))

    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(len(header) + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
