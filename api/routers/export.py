"""Export API router for generating PDF/Excel/CSV reports of health screening data.

Sync port — uses load_district_data() from data_adapter (psycopg2).
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from data.facts import DCODE_TO_ZONE as ZONE_MAPPING, ZONE_NAMES_EN as ZONE_NAMES
from services.data_adapter import load_district_data

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _thai_timestamp() -> str:
    """Return a Thai-locale-friendly timestamp string."""
    now = datetime.now(timezone.utc)
    return now.strftime("%Y-%m-%d %H:%M UTC")


def _district_rows(district: dict[str, Any]) -> list[dict[str, str]]:
    """Flatten a single district's data into tabular rows."""
    rows: list[dict[str, str]] = []
    dcode = district["dcode"]
    name_th = district["name_th"]
    name_en = district["name_en"]
    total = district["total_screened"]

    for disease_key, disease_data in district["diseases"].items():
        for ind_key, ind_data in disease_data["indicators"].items():
            rows.append({
                "district_code": dcode,
                "district_th": name_th,
                "district_en": name_en,
                "total_screened": str(total),
                "disease": disease_key,
                "disease_th": disease_data["name"],
                "disease_en": disease_data["name_en"],
                "pct_at_risk": str(disease_data["pct_at_risk"]),
                "indicator": ind_key,
                "indicator_label": ind_data["label"],
                "unit": ind_data["unit"],
                "mean": str(ind_data.get("mean", "")),
                "cutoff": str(ind_data.get("cutoff", "")),
                "pct_above_cutoff": str(ind_data.get("pct_above_cutoff", "")),
                "count_above": str(ind_data.get("count_above", "")),
            })
    return rows


CSV_COLUMNS = [
    "district_code",
    "district_th",
    "district_en",
    "total_screened",
    "disease",
    "disease_th",
    "disease_en",
    "pct_at_risk",
    "indicator",
    "indicator_label",
    "unit",
    "mean",
    "cutoff",
    "pct_above_cutoff",
    "count_above",
]


def _rows_to_csv_bytes(rows: list[dict[str, str]]) -> bytes:
    """Convert row dicts to UTF-8 CSV bytes (with BOM for Excel compatibility)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def _rows_to_excel_bytes(rows: list[dict[str, str]], sheet_name: str = "Health Data") -> bytes | None:
    """Try to generate Excel bytes via openpyxl. Returns None if unavailable."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")

    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            value = row.get(col_name, "")
            try:
                value = float(value) if "." in value else int(value)
            except (ValueError, TypeError):
                pass
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 4, 14)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_csv_response(rows: list[dict[str, str]], filename: str) -> StreamingResponse:
    """Return a CSV StreamingResponse."""
    content = _rows_to_csv_bytes(rows)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def _build_excel_or_csv_response(rows: list[dict[str, str]], filename: str, sheet_name: str = "Health Data") -> StreamingResponse:
    """Try Excel first, fall back to CSV."""
    excel_bytes = _rows_to_excel_bytes(rows, sheet_name=sheet_name)
    if excel_bytes is not None:
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    return _build_csv_response(rows, filename)


# ---------------------------------------------------------------------------
# PDF report builder (structured text/JSON -- no reportlab dependency)
# ---------------------------------------------------------------------------

def _build_district_pdf_report(district: dict[str, Any]) -> dict[str, Any]:
    """Build a structured report dict for a district (used as PDF substitute)."""
    report: dict[str, Any] = {
        "report_type": "district_health_screening",
        "generated_at": _thai_timestamp(),
        "district": {
            "code": district["dcode"],
            "name_th": district["name_th"],
            "name_en": district["name_en"],
            "total_screened": district["total_screened"],
        },
        "diseases": {},
    }

    for disease_key, disease_data in district["diseases"].items():
        disease_report: dict[str, Any] = {
            "name_th": disease_data["name"],
            "name_en": disease_data["name_en"],
            "pct_at_risk": disease_data["pct_at_risk"],
            "total_screened": disease_data["total_screened"],
            "indicators": {},
        }
        for ind_key, ind_data in disease_data["indicators"].items():
            disease_report["indicators"][ind_key] = {
                "label": ind_data["label"],
                "unit": ind_data["unit"],
                "mean": ind_data.get("mean"),
                "cutoff": ind_data.get("cutoff"),
                "pct_above_cutoff": ind_data.get("pct_above_cutoff"),
                "count_above": ind_data.get("count_above"),
            }
        report["diseases"][disease_key] = disease_report

    return report


def _build_district_pdf_text(district: dict[str, Any]) -> str:
    """Build a human-readable text report for a district."""
    lines: list[str] = []
    lines.append("=" * 70)
    lines.append(f"รายงานผลการคัดกรองสุขภาพ -- {district['name_th']} ({district['name_en']})")
    lines.append(f"Health Screening Report -- District {district['dcode']}")
    lines.append(f"Generated: {_thai_timestamp()}")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"จำนวนผู้เข้ารับการคัดกรอง (Total Screened): {district['total_screened']:,}")
    lines.append("")

    for disease_key, disease_data in district["diseases"].items():
        lines.append("-" * 50)
        lines.append(f"  {disease_data['name']} ({disease_data['name_en']})")
        lines.append(f"  ร้อยละความเสี่ยง (% At Risk): {disease_data['pct_at_risk']}%")
        lines.append("")
        for ind_key, ind_data in disease_data["indicators"].items():
            cutoff_str = f", cutoff={ind_data['cutoff']}" if ind_data.get("cutoff") else ""
            pct_str = f", {ind_data.get('pct_above_cutoff', 'N/A')}% above cutoff" if ind_data.get("pct_above_cutoff") is not None else ""
            count_val = ind_data.get("count_above")
            count_str = f" ({count_val:,} คน)" if count_val is not None else ""
            mean_val = ind_data.get("mean", "N/A")
            unit_val = ind_data.get("unit", "")
            lines.append(f"    {ind_data['label']} ({ind_key}): mean={mean_val} {unit_val}{cutoff_str}{pct_str}{count_str}")
        lines.append("")

    lines.append("=" * 70)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# District-level endpoints
# ---------------------------------------------------------------------------

@router.get("/district/{dcode}/pdf")
def export_district_pdf(dcode: str):
    """Generate a PDF-style report for a district.

    Returns a text report (Content-Type: text/plain with Thai content).
    """
    data = load_district_data()
    if dcode not in data:
        raise HTTPException(status_code=404, detail=f"District {dcode} not found")

    district = data[dcode]
    text_report = _build_district_pdf_text(district)

    return StreamingResponse(
        io.BytesIO(text_report.encode("utf-8")),
        media_type="text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="report_{dcode}_{district["name_en"].replace(" ", "_")}.txt"',
            "X-Report-Format": "text",
            "X-Report-JSON-Available": "true",
        },
    )


@router.get("/district/{dcode}/pdf/json")
def export_district_pdf_json(dcode: str):
    """Return the structured JSON report for a district (alternative to text PDF)."""
    data = load_district_data()
    if dcode not in data:
        raise HTTPException(status_code=404, detail=f"District {dcode} not found")

    return _build_district_pdf_report(data[dcode])


@router.get("/district/{dcode}/excel")
def export_district_excel(dcode: str):
    """Generate Excel/CSV export for a single district with all diseases and indicators."""
    data = load_district_data()
    if dcode not in data:
        raise HTTPException(status_code=404, detail=f"District {dcode} not found")

    district = data[dcode]
    rows = _district_rows(district)
    filename = f"health_district_{dcode}_{district['name_en'].replace(' ', '_')}"
    return _build_excel_or_csv_response(rows, filename, sheet_name=district["name_en"])


# ---------------------------------------------------------------------------
# Zone-level endpoint
# ---------------------------------------------------------------------------

@router.get("/zone/{zone_code}/excel")
def export_zone_excel(zone_code: str):
    """Generate Excel/CSV export for all districts in a zone."""
    if zone_code not in ZONE_NAMES:
        raise HTTPException(status_code=404, detail=f"Zone {zone_code} not found. Valid: {list(ZONE_NAMES.keys())}")

    data = load_district_data()
    zone_dcodes = [dcode for dcode, zc in ZONE_MAPPING.items() if zc == zone_code]
    rows: list[dict[str, str]] = []
    for dcode in sorted(zone_dcodes):
        if dcode in data:
            rows.extend(_district_rows(data[dcode]))

    if not rows:
        raise HTTPException(status_code=404, detail=f"No data found for zone {zone_code}")

    filename = f"health_zone_{zone_code}_{ZONE_NAMES[zone_code].replace(' ', '_')}"
    return _build_excel_or_csv_response(rows, filename, sheet_name=f"Zone {zone_code}")


# ---------------------------------------------------------------------------
# City-wide endpoint
# ---------------------------------------------------------------------------

@router.get("/city/excel")
def export_city_excel():
    """Generate Excel/CSV export for all 50 Bangkok districts."""
    data = load_district_data()
    rows: list[dict[str, str]] = []
    for dcode in sorted(data.keys()):
        rows.extend(_district_rows(data[dcode]))

    filename = "health_bangkok_all_districts"
    return _build_excel_or_csv_response(rows, filename, sheet_name="Bangkok All Districts")


# ---------------------------------------------------------------------------
# Disease ranking endpoint
# ---------------------------------------------------------------------------

@router.get("/rankings/{disease}/excel")
def export_rankings_excel(disease: str):
    """Generate Excel/CSV export of district rankings for a specific disease."""
    data = load_district_data()

    valid_diseases: set[str] = set()
    for d in data.values():
        valid_diseases.update(d["diseases"].keys())
    if disease not in valid_diseases:
        raise HTTPException(status_code=404, detail=f"Disease '{disease}' not found. Valid: {sorted(valid_diseases)}")

    ranking_columns = [
        "rank", "district_code", "district_th", "district_en",
        "total_screened", "disease", "disease_en", "pct_at_risk",
    ]
    entries: list[dict[str, Any]] = []
    for d in data.values():
        if disease in d["diseases"]:
            entries.append({
                "dcode": d["dcode"],
                "name_th": d["name_th"],
                "name_en": d["name_en"],
                "total_screened": d["total_screened"],
                "pct_at_risk": d["diseases"][disease]["pct_at_risk"],
                "disease_en": d["diseases"][disease]["name_en"],
            })
    entries.sort(key=lambda x: x["pct_at_risk"], reverse=True)

    rows: list[dict[str, str]] = []
    for rank, entry in enumerate(entries, start=1):
        rows.append({
            "rank": str(rank),
            "district_code": entry["dcode"],
            "district_th": entry["name_th"],
            "district_en": entry["name_en"],
            "total_screened": str(entry["total_screened"]),
            "disease": disease,
            "disease_en": entry["disease_en"],
            "pct_at_risk": str(entry["pct_at_risk"]),
        })

    # Build CSV/Excel with ranking-specific columns
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=ranking_columns)
    writer.writeheader()
    writer.writerows(rows)
    csv_content = b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")

    # Try Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = f"{disease} Rankings"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

        for col_idx, col_name in enumerate(ranking_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(ranking_columns, start=1):
                value = row.get(col_name, "")
                try:
                    value = float(value) if "." in value else int(value)
                except (ValueError, TypeError):
                    pass
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, col_name in enumerate(ranking_columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 4, 14)

        out = io.BytesIO()
        wb.save(out)
        filename = f"health_rankings_{disease}"
        return StreamingResponse(
            io.BytesIO(out.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )
    except ImportError:
        pass

    # Fallback to CSV
    filename = f"health_rankings_{disease}"
    return StreamingResponse(
        io.BytesIO(csv_content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )
